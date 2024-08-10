# import json
# import boto3
# import pandas as pd
# from io import BytesIO
# import openpyxl



# def lambda_handler(event, context):
#     # Initialize boto3 client for Glue
#     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
#     glue_client = session.client('glue', region_name='us-east-1')
    
#     # Parameters
#     database_name = "leisure-4037-refreshed-data-without-timestamp-crawler-files"
    
#     # Fetch tables in the specified database
#     response = glue_client.get_tables(DatabaseName=database_name)
#     tables = response['TableList']
#     print("LEN TABLES: ", len(tables))
    
#     # Initialize an empty list to store data from all tables
#     all_data_list = []
    
#     # Loop through each table
#     for table in tables:
#         table_name = table['Name']
#         columns = table['StorageDescriptor']['Columns']
        
#         # Extract column names
#         column_names = [col['Name'] for col in columns]
        
#         # Append table name and its fields to the list
#         all_data_list.append({'FirstColumn': table_name, 'SecondColumn': '\n'.join(column_names)})
    
#     # Create DataFrame from the list
#     all_data = pd.DataFrame(all_data_list)
#     print("All Data: ", all_data)
#     # Create an Excel writer object
#     excel_buffer = BytesIO()
#     with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#         all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
#         # Access the workbook
#         workbook = writer.book
#         worksheet = writer.sheets['Consolidated Data']
        
#         # Format the first column (table name)
#         first_column = openpyxl.utils.cell.get_column_letter(1)
#         for row in worksheet.iter_rows(min_row=2, max_row=len(all_data) + 1, min_col=1, max_col=1):
#             for cell in row:
#                 cell.alignment = openpyxl.styles.Alignment(horizontal='center')
#                 cell.font = openpyxl.styles.Font(bold=True)
#                 cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
#         # Format the second column (fields)
#         second_column = openpyxl.utils.cell.get_column_letter(2)
#         for row in worksheet.iter_rows(min_row=2, max_row=len(all_data) + 1, min_col=2, max_col=2):
#             for cell in row:
#                 cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    
#     # Seek to the beginning of the stream
#     excel_buffer.seek(0)
#     local_filename = f'{database_name}_consolidated.xlsx'
#     with open(local_filename, 'wb') as f:
#         f.write(excel_buffer.getvalue())
    
#     # Save the Excel file to S3
#     # destination_access_key_id = 'ASIAXBY2UIXEJ7NX7OAR'
#     # destination_secret_access_key = 'dZ/4uZjcGyCuAfqTC/T7aOnaSbnGCCh+7QvDfF2W'
#     # destination_session_token = 'IQoJb3JpZ2luX2VjEJH//////////wEaCXVzLWVhc3QtMSJIMEYCIQCzT1dlcBQV81EU53kP81HrfAElMqWd98hV+uZe27+GowIhAL51x/k6eY9vjNcYk46W1J9DUQ960Y5IDklxc2Il8BtlKvsCCBoQABoMNDg0ODUwMjg4MDcyIgzV08aF5XY5kcndc8Mq2AL9jr1Wb8/JWjnQhHVLyyAXSeOzLpOSiIRfd/2giNIzP8XWvZxwWMkCmvHucsTYmwF9dzcyeCwRICwTUjyrCSa+KWhBXEBEmEjMwd5ClblMn4I0j8CONthKyPMDQ0SidsDJTr4QhaH/9/rU/OOKsJy+oijrPvIeZ20k1chvLW5sFEz6PqTcsfdhWyT5IYoVtTInwS8VrHZfnnvsIIknjcMCHCrH8Eygy44Gq91BbwhNo8KxetHHeZpgIIzAvh7k1B4Ozi+VPgjDgnpcZk5xTnjH8M/k8WwqtRx0IPB5ivIW+0HnkqeqOT5TENEebOEJN7ndD9MGJotyBkhL9kxOcCWb4HcUtXnpoTPLR7+RHTPJSAhphq1/a76yqrEws0s3IHw8QB922KixvwBfJBKZZz0uKxKT4VuK4kMZLP6rgkbyp+mrCdtFO+zL6NN4c2vuwZlcCTA/BOOBajDiu92yBjqmAW8cqThCAYH4NkBn2RRESVxZmPmPlsoheTVUauoi7wUfPThxEPUweagbCVLt6Dn4mZo5PzLfS6KnhtaaMBkvEMr5w2Pm3H29XevJzbGXIKObBIiJw8ag5xiaUv38XPVMPlbYvYMRbzHzmOnl3YSBAG7idGYeyx/0MhcNWB7xmobxiCAFds401WSZ5kY6KE01seP2Towu+9qkRZMfOCexp2gIvCNO1cg='

#     # s3_client = boto3.client(
#     # 's3',
#     # aws_access_key_id=destination_access_key_id,
#     # aws_secret_access_key=destination_secret_access_key,
#     # aws_session_token=destination_session_token
#     # )
#     # Assume the role for the source bucket
#     # source_credentials = assume_role(source_role_arn)

#     # # Create S3 clients using the temporary credentials for the source and fixed credentials for the destination
#     # source_s3_client = boto3.client(
#     #     's3',
#     #     aws_access_key_id=source_credentials['AccessKeyId'],
#     #     aws_secret_access_key=source_credentials['SecretAccessKey'],
#     #     aws_session_token=source_credentials['SessionToken']
#     # )
#     # s3_client = boto3.client('s3')
#     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
    
#     # Use the session to create a client for S3
#     s3_client = session.client('s3')
    
#     s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
#     s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
#     s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
#     print("files save to s3")
#     return {
#         'statusCode': 200,
#         'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
#     }


# lambda_handler('','')




import json
import boto3
import pandas as pd
from io import BytesIO
import openpyxl

def lambda_handler(event, context):
    # Initialize boto3 client for Glue
    session = boto3.Session(profile_name='PowerUserAccess-484850288072')
    glue_client = session.client('glue', region_name='us-east-1')
    
    # Parameters
    database_name = "mmlist-6554"
    
    # Fetch tables in the specified database with pagination
    all_tables = []
    paginator = glue_client.get_paginator('get_tables')
    for page in paginator.paginate(DatabaseName=database_name):
        all_tables.extend(page['TableList'])
    
    print("LEN TABLES: ", len(all_tables))
    
    # Initialize an empty list to store data from all tables
    all_data_list = []
    
    # Loop through each table
    for table in all_tables:
        table_name = table['Name']
        columns = table['StorageDescriptor']['Columns']
        
        # Extract column names
        column_names = [col['Name'] for col in columns]
        
        # Append table name and its fields to the list
        all_data_list.append({'FirstColumn': table_name, 'SecondColumn': '\n'.join(column_names)})
    
    # Create DataFrame from the list
    all_data = pd.DataFrame(all_data_list)
    print("All Data: ", all_data)
    
    # Create an Excel writer object
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
        # Access the workbook
        workbook = writer.book
        worksheet = writer.sheets['Consolidated Data']
        
        # Format the first column (table name)
        first_column = openpyxl.utils.cell.get_column_letter(1)
        for row in worksheet.iter_rows(min_row=2, max_row=len(all_data) + 1, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        # Format the second column (fields)
        second_column = openpyxl.utils.cell.get_column_letter(2)
        for row in worksheet.iter_rows(min_row=2, max_row=len(all_data) + 1, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    
    # Seek to the beginning of the stream
    excel_buffer.seek(0)
    local_filename = f'{database_name}-consolidated.xlsx'
    with open(local_filename, 'wb') as f:
        f.write(excel_buffer.getvalue())
    
    # Save the Excel file to S3
    session = boto3.Session(profile_name='PowerUserAccess-484850288072')
    s3_client = session.client('s3')
    
    s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
    s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
    s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
    print("files save to s3")
    
    return {
        'statusCode': 200,
        'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
    }

lambda_handler('', '')




# this is the lambda for beeline consolodated