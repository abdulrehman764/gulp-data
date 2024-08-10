# # import json
# # import boto3
# # import pandas as pd
# # from io import BytesIO
# # import openpyxl

# # def lambda_handler(event, context):
# #     # Initialize boto3 clients for Glue and Athena
# #     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
# #     glue_client = session.client('glue', region_name='us-east-1')
# #     athena_client = session.client('athena', region_name='us-east-1')
    
# #     # Parameters
# #     database_name = "mmlist-6554"
# #     s3_output = 's3://ar-athena-query-results-bucket/'  # Replace with your Athena query results bucket
    
# #     # Fetch tables in the specified database with pagination
# #     all_tables = []
# #     paginator = glue_client.get_paginator('get_tables')
# #     for page in paginator.paginate(DatabaseName=database_name):
# #         all_tables.extend(page['TableList'])
    
# #     print("LEN TABLES: ", len(all_tables))
    
# #     # Initialize a dictionary to store data from all tables
# #     all_data_dict = {}
# #     column_count = {}
# #     table_sizes = {}

# #     # Loop through each table
# #     for table in all_tables:
# #         table_name = table['Name']
# #         columns = table['StorageDescriptor']['Columns']
        
# #         # Extract column names
# #         column_names = [col['Name'] for col in columns]
        
# #         # Count occurrences of each column name
# #         for col in column_names:
# #             if col in column_count:
# #                 column_count[col] += 1
# #             else:
# #                 column_count[col] = 1
        
# #         # Append table name and its fields to the dictionary
# #         all_data_dict[table_name] = column_names
        
# #         # Run Athena query to get table size
# #         query = f'SELECT COUNT(*) FROM "{database_name}"."{table_name}"'
# #         response = athena_client.start_query_execution(
# #             QueryString=query,
# #             QueryExecutionContext={'Database': database_name}
# #             ,ResultConfiguration={'OutputLocation': s3_output}
# #         )
# #         query_execution_id = response['QueryExecutionId']
        
# #         # Get query results
# #         result = athena_client.get_query_results(QueryExecutionId=query_execution_id)
# #         table_size = int(result['ResultSet']['Rows'][1]['Data'][0]['VarCharValue'])
# #         table_sizes[table_name] = table_size

# #     # Identify columns that occur more than once
# #     repeated_columns = {col for col, count in column_count.items() if count > 1}

# #     # Sort the tables by number of columns in descending order
# #     sorted_tables = sorted(all_data_dict.items(), key=lambda x: len(x[1]), reverse=True)
# #     sorted_data_dict = {table: columns for table, columns in sorted_tables}

# #     # Create a DataFrame from the sorted dictionary and table sizes
# #     all_data = pd.DataFrame.from_dict(sorted_data_dict, orient='index').transpose()
# #     all_data = all_data.append(pd.Series(name='Table Size', dtype=int))
# #     for table_name in table_sizes:
# #         all_data.at['Table Size', table_name] = table_sizes[table_name]
# #     print("All Data: ", all_data)
    
# #     # Create an Excel writer object
# #     excel_buffer = BytesIO()
# #     with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
# #         all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
# #         # Access the workbook
# #         workbook = writer.book
# #         worksheet = writer.sheets['Consolidated Data']
        
# #         # Insert an empty row after the header and before attributes
# #         worksheet.insert_rows(2)
        
# #         # Write table sizes in the second row
# #         for col_num, table_name in enumerate(all_data.columns, 1):
# #             worksheet.cell(row=2, column=col_num).value = table_sizes.get(table_name, 'N/A')
# #             worksheet.cell(row=2, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
# #             worksheet.cell(row=2, column=col_num).font = openpyxl.styles.Font(bold=True)

# #         # Format the first row (header)
# #         for cell in worksheet[1]:
# #             cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
# #             cell.font = openpyxl.styles.Font(bold=True)

# #         # Format the rest of the cells and change the font color for repeated columns
# #         for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
# #             for cell in row:
# #                 cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
# #                 if cell.value in repeated_columns:
# #                     cell.font = openpyxl.styles.Font(color='FF0000')  # Red color for repeated columns
    
# #     # Seek to the beginning of the stream
# #     excel_buffer.seek(0)
# #     local_filename = f'{database_name}-consolidated.xlsx'
# #     with open(local_filename, 'wb') as f:
# #         f.write(excel_buffer.getvalue())
    
# #     # Save the Excel file to S3
# #     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
# #     s3_client = session.client('s3')
    
# #     s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
# #     s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
# #     s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
# #     print("Files saved to S3")
    
# #     return {
# #         'statusCode': 200,
# #         'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
# #     }

# # lambda_handler('', '')



# # import json
# # import boto3
# # import pandas as pd
# # from io import BytesIO
# # import openpyxl
# # import time

# # def lambda_handler(event, context):
# #     # Initialize boto3 clients for Glue and Athena
# #     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
# #     glue_client = session.client('glue', region_name='us-east-1')
# #     athena_client = session.client('athena', region_name='us-east-1')
    
# #     # Parameters
# #     database_name = "mmlist-6554"
# #     s3_output = 's3://ar-athena-query-results-bucket/'  # Replace with your Athena query results bucket
    
# #     # Fetch tables in the specified database with pagination
# #     all_tables = []
# #     paginator = glue_client.get_paginator('get_tables')
# #     for page in paginator.paginate(DatabaseName=database_name):
# #         all_tables.extend(page['TableList'])
    
# #     print("LEN TABLES: ", len(all_tables))
    
# #     # Initialize a dictionary to store data from all tables
# #     all_data_dict = {}
# #     column_count = {}
# #     table_sizes = {}

# #     # Loop through each table
# #     for table in all_tables:
# #         table_name = table['Name']
# #         columns = table['StorageDescriptor']['Columns']
        
# #         # Extract column names
# #         column_names = [col['Name'] for col in columns]
        
# #         # Count occurrences of each column name
# #         for col in column_names:
# #             if col in column_count:
# #                 column_count[col] += 1
# #             else:
# #                 column_count[col] = 1
        
# #         # Append table name and its fields to the dictionary
# #         all_data_dict[table_name] = column_names
        
# #         # Run Athena query to get table size
# #         query = f'SELECT DISTINCT COUNT(*) FROM "{database_name}"."{table_name}"'
# #         response = athena_client.start_query_execution(
# #             QueryString=query,
# #             QueryExecutionContext={'Database': database_name},
# #             ResultConfiguration={'OutputLocation': s3_output}
# #         )
# #         query_execution_id = response['QueryExecutionId']
        
# #         # Polling the query execution status until it completes
# #         while True:
# #             query_status = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
# #             status = query_status['QueryExecution']['Status']['State']
# #             if status in ['SUCCEEDED', 'FAILED', 'CANCELLED']:
# #                 break
# #             time.sleep(1)
        
# #         if status == 'SUCCEEDED':
# #             result = athena_client.get_query_results(QueryExecutionId=query_execution_id)
# #             table_size = int(result['ResultSet']['Rows'][1]['Data'][0]['VarCharValue'])
# #             table_sizes[table_name] = table_size
# #         else:
# #             print(f"Query failed for table {table_name}: {status}")
# #             table_sizes[table_name] = 'N/A'

# #     # Identify columns that occur more than once
# #     repeated_columns = {col for col, count in column_count.items() if count > 1}

# #     # Sort the tables by number of columns in descending order
# #     sorted_tables = sorted(all_data_dict.items(), key=lambda x: len(x[1]), reverse=True)
# #     sorted_data_dict = {table: columns for table, columns in sorted_tables}

# #     # Create a DataFrame from the sorted dictionary and table sizes
# #     all_data = pd.DataFrame.from_dict(sorted_data_dict, orient='index').transpose()
# #     # Create a DataFrame for table sizes and concatenate it
# #     table_sizes_df = pd.DataFrame(table_sizes, index=['Table Size'])
# #     all_data = pd.concat([table_sizes_df, all_data], axis=0)

# #     print("All Data: ", all_data)
    
# #     # Create an Excel writer object
# #     excel_buffer = BytesIO()
# #     with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
# #         all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
# #         # Access the workbook
# #         workbook = writer.book
# #         worksheet = writer.sheets['Consolidated Data']
        
# #         # Insert an empty row after the header and before attributes
# #         worksheet.insert_rows(2)
        
# #         # Write table sizes in the second row
# #         for col_num, table_name in enumerate(all_data.columns, 1):
# #             worksheet.cell(row=2, column=col_num).value = table_sizes.get(table_name, 'N/A')
# #             worksheet.cell(row=2, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
# #             worksheet.cell(row=2, column=col_num).font = openpyxl.styles.Font(bold=True)

# #         # Format the first row (header)
# #         for cell in worksheet[1]:
# #             cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
# #             cell.font = openpyxl.styles.Font(bold=True)

# #         # Format the rest of the cells and change the font color for repeated columns
# #         for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
# #             for cell in row:
# #                 cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
# #                 if cell.value in repeated_columns:
# #                     cell.font = openpyxl.styles.Font(color='FF0000')  # Red color for repeated columns
    
# #     # Seek to the beginning of the stream
# #     excel_buffer.seek(0)
# #     local_filename = f'{database_name}-consolidated.xlsx'
# #     with open(local_filename, 'wb') as f:
# #         f.write(excel_buffer.getvalue())
    
# #     # Save the Excel file to S3
# #     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
# #     s3_client = session.client('s3')
    
# #     s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
# #     s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
# #     s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
# #     print("Files saved to S3")
    
# #     return {
# #         'statusCode': 200,
# #         'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
# #     }

# # lambda_handler('', '')







# import json
# import boto3
# import pandas as pd
# from io import BytesIO
# import openpyxl
# import time

# def lambda_handler(event, context):
#     print("Starting Lambda Execution...")
#     # Initialize boto3 clients for Glue and Athena
#     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
#     glue_client = session.client('glue', region_name='us-east-1')
#     athena_client = session.client('athena', region_name='us-east-1')
    
#     # Parameters
#     # database_name = "mmlist-6554"
#     # database_name = "leisure-4037-refreshed-data-without-timestamp-crawler-files"
#     # database_name = "beeline-4134-processed-crawl-test-final"
#     # database_name = "community-finance-database"
#     # database_name = "uva-1331-updated"
#     database_name = "fortuna-database"
#     s3_output = 's3://ar-athena-query-results-bucket/'  # Replace with your Athena query results bucket
    
    

    
#     print(f"Fetching tables in database: {database_name}...")
#     # Fetch tables in the specified database with pagination
#     all_tables = []
#     paginator = glue_client.get_paginator('get_tables')
#     for page in paginator.paginate(DatabaseName=database_name):
#         all_tables.extend(page['TableList'])
    
#     print("Number of tables found:", len(all_tables))
    
#     # Initialize a dictionary to store data from all tables
#     all_data_dict = {}
#     column_count = {}
#     table_sizes = {}

#     # Loop through each table
#     for table in all_tables:
#         table_name = table['Name']
#         columns = table['StorageDescriptor']['Columns']
        
#         # Extract column names
#         column_names = [col['Name'] for col in columns]
        
#         # Count occurrences of each column name
#         for col in column_names:
#             if col in column_count:
#                 column_count[col] += 1
#             else:
#                 column_count[col] = 1
        
#         # Append table name and its fields to the dictionary
#         all_data_dict[table_name] = column_names
        
#         # Run Athena query to get table size
#         print(f"Fetching size for table: {table_name}...")
#         query = f'SELECT DISTINCT COUNT(*) FROM "{database_name}"."{table_name}"'
#         response = athena_client.start_query_execution(
#             QueryString=query,
#             QueryExecutionContext={'Database': database_name},
#             ResultConfiguration={'OutputLocation': s3_output}
#         )
#         query_execution_id = response['QueryExecutionId']
        
#         # Polling the query execution status until it completes
#         while True:
#             query_status = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
#             status = query_status['QueryExecution']['Status']['State']
#             print(f"Query status for {table_name}: {status}")
#             if status in ['SUCCEEDED', 'FAILED', 'CANCELLED']:
#                 break
#             time.sleep(1)
        
#         if status == 'SUCCEEDED':
#             result = athena_client.get_query_results(QueryExecutionId=query_execution_id)
#             table_size = int(result['ResultSet']['Rows'][1]['Data'][0]['VarCharValue'])
#             table_sizes[table_name] = table_size
#         else:
#             print(f"Query failed for table {table_name}: {status}")
#             table_sizes[table_name] = 'N/A'

#     # Identify columns that occur more than once
#     repeated_columns = {col for col, count in column_count.items() if count > 1}

#     # Sort the tables by number of columns in descending order
#     sorted_tables = sorted(all_data_dict.items(), key=lambda x: len(x[1]), reverse=True)
#     sorted_data_dict = {table: columns for table, columns in sorted_tables}

#     # Create a DataFrame from the sorted dictionary
#     all_data = pd.DataFrame.from_dict(sorted_data_dict, orient='index').transpose()

#     # Create a DataFrame for table sizes
#     table_sizes_df = pd.DataFrame(table_sizes.values(), index=table_sizes.keys(), columns=['Table Size']).transpose()

#     # Concatenate table sizes DataFrame at the bottom of all_data
#     all_data = pd.concat([all_data, table_sizes_df])

#     print("All Data: ", all_data)
    
#     # Create an Excel writer object
#     excel_buffer = BytesIO()
#     with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#         all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
#         # Access the workbook
#         workbook = writer.book
#         worksheet = writer.sheets['Consolidated Data']
        
#         # Insert an empty row after the header and before attributes
#         worksheet.insert_rows(2)
        
#         # Write table sizes in the second row
#         for col_num, table_name in enumerate(all_data.columns, 1):
#             worksheet.cell(row=2, column=col_num).value = table_sizes.get(table_name, 'N/A')
#             worksheet.cell(row=2, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
#             worksheet.cell(row=2, column=col_num).font = openpyxl.styles.Font(bold=True)

#         # Format the first row (header)
#         for cell in worksheet[1]:
#             cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
#             cell.font = openpyxl.styles.Font(bold=True)

#         # Format the rest of the cells and change the font color for repeated columns
#         for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
#             for cell in row:
#                 cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
#                 if cell.value in repeated_columns:
#                     cell.font = openpyxl.styles.Font(color='FF0000')  # Red color for repeated columns
    
#     # Seek to the beginning of the stream
#     excel_buffer.seek(0)
#     local_filename = f'{database_name}-consolidated.xlsx'
#     with open(local_filename, 'wb') as f:
#         f.write(excel_buffer.getvalue())
    
#     # Save the Excel file to S3
#     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
#     s3_client = session.client('s3')
    
#     s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
#     s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
#     s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
#     print("Files saved to S3")
    
#     return {
#         'statusCode': 200,
#         'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
#     }

# lambda_handler('', '')






























# import json
# import boto3
# import pandas as pd
# from io import BytesIO
# import openpyxl
# import time

# def lambda_handler(event, context):
#     print("Starting Lambda Execution...")
#     # Initialize boto3 clients for Glue and Athena
#     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
#     glue_client = session.client('glue', region_name='us-east-1')
#     athena_client = session.client('athena', region_name='us-east-1')
    
#     # Parameters
#     database_name = "fortuna-database"
#         # Parameters
#     database_name = "community-finance-database"
    
#     s3_output = 's3://ar-athena-query-results-bucket/'  # Replace with your Athena query results bucket
    
#     print(f"Fetching tables in database: {database_name}...")
#     # Fetch tables in the specified database with pagination
#     all_tables = []
#     paginator = glue_client.get_paginator('get_tables')
#     for page in paginator.paginate(DatabaseName=database_name):
#         all_tables.extend(page['TableList'])
    
#     print("Number of tables found:", len(all_tables))
    
#     # Initialize a dictionary to store data from all tables
#     all_data_dict = {}
#     column_count = {}
#     table_sizes = {}
#     create_table_queries = {}

#     # Loop through each table
#     for table in all_tables:
#         table_name = table['Name']
#         columns = table['StorageDescriptor']['Columns']
        
#         # Extract column names
#         column_names = [col['Name'] for col in columns]
        
#         # Count occurrences of each column name
#         for col in column_names:
#             if col in column_count:
#                 column_count[col] += 1
#             else:
#                 column_count[col] = 1
        
#         # Append table name and its fields to the dictionary
#         all_data_dict[table_name] = column_names
        
#         # Run Athena query to get table size
#         print(f"Fetching size for table: {table_name}...")
#         query = f'SELECT DISTINCT COUNT(*) FROM "{database_name}"."{table_name}"'
#         response = athena_client.start_query_execution(
#             QueryString=query,
#             QueryExecutionContext={'Database': database_name},
#             ResultConfiguration={'OutputLocation': s3_output}
#         )
#         query_execution_id = response['QueryExecutionId']
        
#         # Polling the query execution status until it completes
#         while True:
#             query_status = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
#             status = query_status['QueryExecution']['Status']['State']
#             print(f"Query status for {table_name}: {status}")
#             if status in ['SUCCEEDED', 'FAILED', 'CANCELLED']:
#                 break
#             time.sleep(1)
        
#         if status == 'SUCCEEDED':
#             result = athena_client.get_query_results(QueryExecutionId=query_execution_id)
#             if len(result['ResultSet']['Rows']) > 1:
#                 table_size = int(result['ResultSet']['Rows'][1]['Data'][0]['VarCharValue'])
#                 table_sizes[table_name] = table_size
#             else:
#                 print(f"No result returned for table size of {table_name}")
#                 table_sizes[table_name] = 'N/A'
#         else:
#             print(f"Query failed for table {table_name}: {status}")
#             table_sizes[table_name] = 'N/A'

#         # Run Athena query to get SHOW CREATE TABLE statement
#         print(f"Fetching CREATE TABLE statement for: {table_name}...")
#         create_table_query = f'SHOW CREATE TABLE `{database_name}.{table_name}`'
#         response = athena_client.start_query_execution(
#             QueryString=create_table_query,
#             QueryExecutionContext={'Database': database_name},
#             ResultConfiguration={'OutputLocation': s3_output}
#         )
#         query_execution_id = response['QueryExecutionId']
        
#         # Polling the query execution status until it completes
#         while True:
#             query_status = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
#             status = query_status['QueryExecution']['Status']['State']
#             print(f"CREATE TABLE Query status for {table_name}: {status}")
#             if status in ['SUCCEEDED', 'FAILED', 'CANCELLED']:
#                 break
#             time.sleep(1)
        
#         if status == 'SUCCEEDED':
#             result = athena_client.get_query_results(QueryExecutionId=query_execution_id)
#             if len(result['ResultSet']['Rows']) > 1:
#                 create_table_statement = result['ResultSet']['Rows'][1]['Data'][0]['VarCharValue']
#                 create_table_queries[table_name] = create_table_statement
#             else:
#                 print(f"No CREATE TABLE statement returned for {table_name}")
#                 create_table_queries[table_name] = 'N/A'
#         else:
#             print(f"CREATE TABLE query failed for table {table_name}: {status}")
#             create_table_queries[table_name] = 'N/A'

#     # Identify columns that occur more than once
#     repeated_columns = {col for col, count in column_count.items() if count > 1}

#     # Sort the tables by number of columns in descending order
#     sorted_tables = sorted(all_data_dict.items(), key=lambda x: len(x[1]), reverse=True)
#     sorted_data_dict = {table: columns for table, columns in sorted_tables}

#     # Create a DataFrame from the sorted dictionary
#     all_data = pd.DataFrame.from_dict(sorted_data_dict, orient='index').transpose()

#     # Create a DataFrame for table sizes
#     table_sizes_df = pd.DataFrame(table_sizes.values(), index=table_sizes.keys(), columns=['Table Size']).transpose()

#     # Concatenate table sizes DataFrame at the bottom of all_data
#     all_data = pd.concat([all_data, table_sizes_df])

#     print("All Data: ", all_data)
    
#     # Create an Excel writer object
#     excel_buffer = BytesIO()
#     with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#         all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
#         # Access the workbook
#         workbook = writer.book
#         worksheet = writer.sheets['Consolidated Data']
        
#         # Insert an empty row after the header and before attributes
#         worksheet.insert_rows(2)
        
#         # Write table sizes in the second row
#         for col_num, table_name in enumerate(all_data.columns, 1):
#             worksheet.cell(row=2, column=col_num).value = table_sizes.get(table_name, 'N/A')
#             worksheet.cell(row=2, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
#             worksheet.cell(row=2, column=col_num).font = openpyxl.styles.Font(bold=True)

#         # Format the first row (header)
#         for cell in worksheet[1]:
#             cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
#             cell.font = openpyxl.styles.Font(bold=True)

#         # Format the rest of the cells and change the font color for repeated columns
#         for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
#             for cell in row:
#                 cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
#                 if cell.value in repeated_columns:
#                     cell.font = openpyxl.styles.Font(color='FF0000')  # Red color for repeated columns

#         # Create a new sheet for the CREATE TABLE statements
#         create_table_df = pd.DataFrame.from_dict(create_table_queries, orient='index', columns=['Create Table Statement'])
#         create_table_df.to_excel(writer, sheet_name='Create Table Statements')
    
#     # Seek to the beginning of the stream
#     excel_buffer.seek(0)
#     local_filename = f'{database_name}-consolidated.xlsx'
#     with open(local_filename, 'wb') as f:
#         f.write(excel_buffer.getvalue())
    
#     # Save the Excel file to S3
#     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
#     s3_client = session.client('s3')
    
#     s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
#     s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
#     s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
#     print("Files saved to S3")
    
#     return {
#         'statusCode': 200,
#         'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
#     }

# lambda_handler('', '')











import json
import boto3
import pandas as pd
from io import BytesIO
import openpyxl
import time

def lambda_handler(event, context):
    print("Starting Lambda Execution...")
    # Initialize boto3 clients for Glue and Athena
    session = boto3.Session(profile_name='PowerUserAccess-484850288072')
    glue_client = session.client('glue', region_name='us-east-1')
    athena_client = session.client('athena', region_name='us-east-1')
    
    # Parameters
    database_name = "fortuna-database"
    database_name = "community-finance-database"

    database_name = "leisure-4037-refreshed-data-without-timestamp-crawler-files"
    database_name = "uva-1331"
    s3_output = 's3://ar-athena-query-results-bucket/'  # Replace with your Athena query results bucket
    
    print(f"Fetching tables in database: {database_name}...")
    # Fetch tables in the specified database with pagination
    all_tables = []
    paginator = glue_client.get_paginator('get_tables')
    for page in paginator.paginate(DatabaseName=database_name):
        all_tables.extend(page['TableList'])
    
    print("Number of tables found:", len(all_tables))
    
    # Initialize a dictionary to store data from all tables
    all_data_dict = {}
    column_count = {}
    table_sizes = {}
    create_statements = {}

    # Loop through each table
    for table in all_tables:
        table_name = table['Name']
        columns = table['StorageDescriptor']['Columns']
        
        # Extract column names
        column_names = [col['Name'] for col in columns]
        
        # Count occurrences of each column name
        for col in column_names:
            if col in column_count:
                column_count[col] += 1
            else:
                column_count[col] = 1
        
        # Append table name and its fields to the dictionary
        all_data_dict[table_name] = column_names
        
        # Run Athena query to get table size
        print(f"Fetching size for table: {table_name}...")
        query = f'SELECT DISTINCT COUNT(*) FROM "{database_name}"."{table_name}"'
        response = athena_client.start_query_execution(
            QueryString=query,
            QueryExecutionContext={'Database': database_name},
            ResultConfiguration={'OutputLocation': s3_output}
        )
        query_execution_id = response['QueryExecutionId']
        
        # Polling the query execution status until it completes
        while True:
            query_status = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
            status = query_status['QueryExecution']['Status']['State']
            print(f"Query status for {table_name}: {status}")
            if status in ['SUCCEEDED', 'FAILED', 'CANCELLED']:
                break
            time.sleep(1)
        
        if status == 'SUCCEEDED':
            result = athena_client.get_query_results(QueryExecutionId=query_execution_id)
            table_size = int(result['ResultSet']['Rows'][1]['Data'][0]['VarCharValue'])
            table_sizes[table_name] = table_size
        else:
            print(f"Query failed for table {table_name}: {status}")
            table_sizes[table_name] = 'N/A'

        # Run Athena query to get the CREATE TABLE statement
        print(f"Fetching CREATE TABLE statement for table: {table_name}...")
        query = f'SHOW CREATE TABLE `{database_name}.{table_name}`'
        response = athena_client.start_query_execution(
            QueryString=query,
            QueryExecutionContext={'Database': database_name},
            ResultConfiguration={'OutputLocation': s3_output}
        )
        query_execution_id = response['QueryExecutionId']

        # Polling the query execution status until it completes
        while True:
            query_status = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
            status = query_status['QueryExecution']['Status']['State']
            print(f"Query status for {table_name} (SHOW CREATE TABLE): {status}")
            if status in ['SUCCEEDED', 'FAILED', 'CANCELLED']:
                break
            time.sleep(1)
        
        if status == 'SUCCEEDED':
            result = athena_client.get_query_results(QueryExecutionId=query_execution_id)
            # Concatenate all parts of the CREATE TABLE statement
            create_statement = "".join([row['Data'][0]['VarCharValue'] for row in result['ResultSet']['Rows']])
            create_statements[table_name] = create_statement
        else:
            print(f"SHOW CREATE TABLE failed for table {table_name}: {status}")
            create_statements[table_name] = 'N/A'

    # Identify columns that occur more than once
    repeated_columns = {col for col, count in column_count.items() if count > 1}

    # Sort the tables by number of columns in descending order
    sorted_tables = sorted(all_data_dict.items(), key=lambda x: len(x[1]), reverse=True)
    sorted_data_dict = {table: columns for table, columns in sorted_tables}

    # Create a DataFrame from the sorted dictionary
    all_data = pd.DataFrame.from_dict(sorted_data_dict, orient='index').transpose()

    # Create a DataFrame for table sizes and CREATE TABLE statements
    table_sizes_df = pd.DataFrame(table_sizes.values(), index=table_sizes.keys(), columns=['Table Size']).transpose()
    create_statements_df = pd.DataFrame(create_statements.values(), index=create_statements.keys(), columns=['CREATE TABLE']).transpose()

    # Concatenate table sizes and CREATE TABLE statements DataFrame at the bottom of all_data
    all_data = pd.concat([all_data, table_sizes_df, create_statements_df])

    print("All Data: ", all_data)
    
    # Create an Excel writer object
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
        # Access the workbook
        workbook = writer.book
        worksheet = writer.sheets['Consolidated Data']
        
        # Insert an empty row after the header and before attributes
        worksheet.insert_rows(2)
        
        # Write table sizes and CREATE TABLE statements in the second row
        for col_num, table_name in enumerate(all_data.columns, 1):
            worksheet.cell(row=2, column=col_num).value = table_sizes.get(table_name, 'N/A')
            worksheet.cell(row=2, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
            worksheet.cell(row=2, column=col_num).font = openpyxl.styles.Font(bold=True)

        # Format the first row (header)
        for cell in worksheet[1]:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
            cell.font = openpyxl.styles.Font(bold=True)

        # Format the rest of the cells and change the font color for repeated columns
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                if cell.value in repeated_columns:
                    cell.font = openpyxl.styles.Font(color='FF0000')  # Red color for repeated columns
    
    # Seek to the beginning of the stream
    excel_buffer.seek(0)
    local_filename = f'{database_name}-consolidated.xlsx'
    with open(local_filename, 'wb') as f:
        f.write(excel_buffer.getvalue())
    
    # Save the Excel file to S3
    session = boto3.Session(profile_name='PowerUserAccess-484850288072')
    s3_client = session.client('s3')
    
    s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
    s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
    s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
    print("Files saved to S3")
    
    return {
        'statusCode': 200,
        'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
    }

lambda_handler('', '')
