import json
import boto3
import pandas as pd
from io import BytesIO
import openpyxl
import time

def lambda_handler(event, context):
    print("Starting Lambda Execution...")
    # Initialize boto3 clients for Glue and Athena
    session = boto3.Session(profile_name='PowerUserAccess-484850288072')
    glue_client = session.client('glue', region_name='us-east-1')
    athena_client = session.client('athena', region_name='us-east-1')
    
    # Parameters
    database_name = "mmlist-6554"
    s3_output = 's3://ar-athena-query-results-bucket/'  # Replace with your Athena query results bucket
    
    print(f"Fetching tables in database: {database_name}...")
    # Fetch tables in the specified database with pagination
    all_tables = []
    paginator = glue_client.get_paginator('get_tables')
    for page in paginator.paginate(DatabaseName=database_name):
        all_tables.extend(page['TableList'])
    
    print("Number of tables found:", len(all_tables))
    
    # Initialize dictionaries to store data
    table_sizes = {}
    all_data_dict = {}

    # Loop through each table
    for table in all_tables:
        table_name = table['Name']
        columns = table['StorageDescriptor']['Columns']
        
        # Extract column names
        column_names = [col['Name'] for col in columns]
        
        # Append table name and its fields to the dictionary
        all_data_dict[table_name] = column_names
        
        # Run Athena query to get table size
        print(f"Fetching size for table: {table_name}...")
        query = f'SELECT DISTINCT COUNT(*) FROM "{database_name}"."{table_name}"'
        response = athena_client.start_query_execution(
            QueryString=query,
            QueryExecutionContext={'Database': database_name},
            ResultConfiguration={'OutputLocation': s3_output}
        )
        query_execution_id = response['QueryExecutionId']
        
        # Polling the query execution status until it completes
        while True:
            query_status = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
            status = query_status['QueryExecution']['Status']['State']
            print(f"Query status for {table_name}: {status}")
            if status in ['SUCCEEDED', 'FAILED', 'CANCELLED']:
                break
            time.sleep(1)
        
        if status == 'SUCCEEDED':
            result = athena_client.get_query_results(QueryExecutionId=query_execution_id)
            table_size = int(result['ResultSet']['Rows'][1]['Data'][0]['VarCharValue'])
            table_sizes[table_name] = table_size
        else:
            print(f"Query failed for table {table_name}: {status}")
            table_sizes[table_name] = 'N/A'

    # Sort the tables by number of columns in descending order
    sorted_tables = sorted(all_data_dict.items(), key=lambda x: len(x[1]), reverse=True)
    sorted_data_dict = {table: columns for table, columns in sorted_tables}

    # Create a DataFrame from the sorted dictionary
    all_data = pd.DataFrame.from_dict(sorted_data_dict, orient='index').transpose()

    # Create a DataFrame for table sizes
    table_sizes_df = pd.DataFrame(table_sizes.values(), index=table_sizes.keys(), columns=['Sample Size']).transpose()

    # Concatenate table sizes DataFrame at the bottom of all_data
    all_data = pd.concat([table_sizes_df, all_data])

    print("All Data: ", all_data)
    
    # Create an Excel writer object
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        all_data.to_excel(writer, index=True, sheet_name='Consolidated Data', startcol=2)
        
        # Access the workbook
        workbook = writer.book
        worksheet = writer.sheets['Consolidated Data']
        
        # Format the first row (header)
        for cell in worksheet[1]:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
            cell.font = openpyxl.styles.Font(bold=True)

        # Format the rest of the cells and change the font color for repeated columns
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    
    # Seek to the beginning of the stream
    excel_buffer.seek(0)
    local_filename = f'{database_name}-consolidated.xlsx'
    with open(local_filename, 'wb') as f:
        f.write(excel_buffer.getvalue())
    
    # Save the Excel file to S3
    s3_client = session.client('s3')
    s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
    s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
    s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
    print("Files saved to S3")
    
    return {
        'statusCode': 200,
        'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
    }

lambda_handler('', '')
