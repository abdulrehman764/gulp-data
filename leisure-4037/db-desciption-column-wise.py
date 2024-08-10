# # import json
# # import boto3
# # import pandas as pd
# # from io import BytesIO
# # import openpyxl

# # def lambda_handler(event, context):
# #     # Initialize boto3 client for Glue
# #     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
# #     glue_client = session.client('glue', region_name='us-east-1')
    
# #     # Parameters
# #     database_name = "mmlist-6554"
    
# #     # Fetch tables in the specified database with pagination
# #     all_tables = []
# #     paginator = glue_client.get_paginator('get_tables')
# #     for page in paginator.paginate(DatabaseName=database_name):
# #         all_tables.extend(page['TableList'])
    
# #     print("LEN TABLES: ", len(all_tables))
    
# #     # Initialize a dictionary to store data from all tables
# #     all_data_dict = {}

# #     # Loop through each table
# #     for table in all_tables:
# #         table_name = table['Name']
# #         columns = table['StorageDescriptor']['Columns']
        
# #         # Extract column names
# #         column_names = [col['Name'] for col in columns]
        
# #         # Append table name and its fields to the dictionary
# #         all_data_dict[table_name] = column_names

# #     # Create a DataFrame from the dictionary
# #     all_data = pd.DataFrame.from_dict(all_data_dict, orient='index').transpose()
# #     print("All Data: ", all_data)
    
# #     # Create an Excel writer object
# #     excel_buffer = BytesIO()
# #     with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
# #         all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
# #         # Access the workbook
# #         workbook = writer.book
# #         worksheet = writer.sheets['Consolidated Data']
        
# #         # Format the first row (header)
# #         for cell in worksheet[1]:
# #             cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
# #             cell.font = openpyxl.styles.Font(bold=True)

# #         # Format the rest of the cells
# #         for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
# #             for cell in row:
# #                 cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    
# #     # Seek to the beginning of the stream
# #     excel_buffer.seek(0)
# #     local_filename = f'{database_name}-consolidated.xlsx'
# #     with open(local_filename, 'wb') as f:
# #         f.write(excel_buffer.getvalue())
    
# #     # Save the Excel file to S3
# #     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
# #     s3_client = session.client('s3')
    
# #     s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
# #     s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
# #     s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
# #     print("Files saved to S3")
    
# #     return {
# #         'statusCode': 200,
# #         'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
# #     }

# # lambda_handler('', '')


# import json
# import boto3
# import pandas as pd
# from io import BytesIO
# import openpyxl

# def lambda_handler(event, context):
#     # Initialize boto3 client for Glue
#     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
#     glue_client = session.client('glue', region_name='us-east-1')
    
#     # Parameters
#     database_name = "mmlist-6554"
    
#     # Fetch tables in the specified database with pagination
#     all_tables = []
#     paginator = glue_client.get_paginator('get_tables')
#     for page in paginator.paginate(DatabaseName=database_name):
#         all_tables.extend(page['TableList'])
    
#     print("LEN TABLES: ", len(all_tables))
    
#     # Initialize a dictionary to store data from all tables
#     all_data_dict = {}

#     # Loop through each table
#     for table in all_tables:
#         table_name = table['Name']
#         columns = table['StorageDescriptor']['Columns']
        
#         # Extract column names
#         column_names = [col['Name'] for col in columns]
        
#         # Append table name and its fields to the dictionary
#         all_data_dict[table_name] = column_names
    
#     # Sort the tables by number of columns in descending order
#     sorted_tables = sorted(all_data_dict.items(), key=lambda x: len(x[1]), reverse=True)
#     sorted_data_dict = {table: columns for table, columns in sorted_tables}

#     # Create a DataFrame from the sorted dictionary
#     all_data = pd.DataFrame.from_dict(sorted_data_dict, orient='index').transpose()
#     print("All Data: ", all_data)
    
#     # Create an Excel writer object
#     excel_buffer = BytesIO()
#     with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#         all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
#         # Access the workbook
#         workbook = writer.book
#         worksheet = writer.sheets['Consolidated Data']
        
#         # Format the first row (header)
#         for cell in worksheet[1]:
#             cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
#             cell.font = openpyxl.styles.Font(bold=True)

#         # Format the rest of the cells
#         for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
#             for cell in row:
#                 cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    
#     # Seek to the beginning of the stream
#     excel_buffer.seek(0)
#     local_filename = f'{database_name}-consolidated.xlsx'
#     with open(local_filename, 'wb') as f:
#         f.write(excel_buffer.getvalue())
    
#     # Save the Excel file to S3
#     session = boto3.Session(profile_name='PowerUserAccess-484850288072')
#     s3_client = session.client('s3')
    
#     s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
#     s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
#     s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
#     print("Files saved to S3")
    
#     return {
#         'statusCode': 200,
#         'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
#     }

# lambda_handler('', '')








import json
import boto3
import pandas as pd
from io import BytesIO
import openpyxl

def lambda_handler(event, context):
    # Initialize boto3 client for Glue
    session = boto3.Session(profile_name='PowerUserAccess-484850288072')
    glue_client = session.client('glue', region_name='us-east-1')
    
    # Parameters
    database_name = "mmlist-6554"
    database_name = "lion-gaming-4116"
    database_name = "community-finance-database"
    
    # Fetch tables in the specified database with pagination
    all_tables = []
    paginator = glue_client.get_paginator('get_tables')
    for page in paginator.paginate(DatabaseName=database_name):
        all_tables.extend(page['TableList'])
    
    print("LEN TABLES: ", len(all_tables))
    
    # Initialize a dictionary to store data from all tables
    all_data_dict = {}
    column_count = {}

    # Loop through each table
    for table in all_tables:
        table_name = table['Name']
        columns = table['StorageDescriptor']['Columns']
        
        # Extract column names
        column_names = [col['Name'] for col in columns]
        
        # Count occurrences of each column name
        for col in column_names:
            if col in column_count:
                column_count[col] += 1
            else:
                column_count[col] = 1
        
        # Append table name and its fields to the dictionary
        all_data_dict[table_name] = column_names
    
    # Identify columns that occur more than once
    repeated_columns = {col for col, count in column_count.items() if count > 1}

    # Sort the tables by number of columns in descending order
    sorted_tables = sorted(all_data_dict.items(), key=lambda x: len(x[1]), reverse=True)
    sorted_data_dict = {table: columns for table, columns in sorted_tables}

    # Create a DataFrame from the sorted dictionary
    all_data = pd.DataFrame.from_dict(sorted_data_dict, orient='index').transpose()
    print("All Data: ", all_data)
    
    # Create an Excel writer object
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        all_data.to_excel(writer, index=False, sheet_name='Consolidated Data')
        
        # Access the workbook
        workbook = writer.book
        worksheet = writer.sheets['Consolidated Data']
        
        # Format the first row (header)
        for cell in worksheet[1]:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
            cell.font = openpyxl.styles.Font(bold=True)

        # Format the rest of the cells and change the font color for repeated columns
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                if cell.value in repeated_columns:
                    cell.font = openpyxl.styles.Font(color='FF0000')  # Red color for repeated columns
    
    # Seek to the beginning of the stream
    excel_buffer.seek(0)
    local_filename = f'{database_name}-consolidated.xlsx'
    with open(local_filename, 'wb') as f:
        f.write(excel_buffer.getvalue())
    
    # Save the Excel file to S3
    session = boto3.Session(profile_name='PowerUserAccess-484850288072')
    s3_client = session.client('s3')
    
    s3_bucket = 'gulp-data-vault-decode'  # Replace with your S3 bucket name
    s3_key = f'{database_name}_consolidated.xlsx'  # Save the file with a unique key
    s3_client.put_object(Bucket=s3_bucket, Key=s3_key, Body=excel_buffer.getvalue())
    print("Files saved to S3")
    
    return {
        'statusCode': 200,
        'body': json.dumps(f'Consolidated file saved to S3 bucket: {s3_bucket}, with key: {s3_key}')
    }

lambda_handler('', '')
